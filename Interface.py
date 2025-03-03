# -*- coding: utf-8 -*-
"""
Created on Wed Dec 18 17:41:26 2024

@author: MaximilianoAlarcon
"""
import re
import sys
import time
from PyQt5.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QPushButton, QLabel, QProgressBar, QFileDialog, QHBoxLayout, QComboBox 
)
from PyQt5.QtCore import Qt, QThread, pyqtSignal, pyqtSlot

from openpyxl import load_workbook
import pandas as pd
import numpy as np


class WorkerThread(QThread):
    progress = pyqtSignal(int)
    finished = pyqtSignal()

    def __init__(self, input_data):
        super().__init__()
        self.input_data = input_data[0]
        self.Formato = input_data[1]
        self.output_data = None

    def Numero(self,value): 
        try:
            return float(str(value).strip())
        except ValueError:
            return value
    
        
    def run(self):
        Total_Registro = []
        ultima_agrup = "&&"
        Total = len(self.input_data)
        
        for ll in range(len(self.input_data)):
            texto = self.input_data[ll]
            if texto[0] != ultima_agrup:
                if ll > 0:
                    Tabla_Final = pd.concat(Agrupacion_Registros)
                    Total_Registro.append(Tabla_Final)

                Agrupacion_Registros = []

            R0 = self.Formato[self.Formato["Registro"] == int(texto[0])]
            columna = ["Original"]+list(R0["Descripcion"].values)[:]
            Data_Procesado = pd.DataFrame(columns=columna)

            Separados = []
            for indice, data in R0.iterrows():
                if data.Descripcion in columna:
                    unidad = texto[int(data.Desde) - 1:int(data.Hasta)]
                    Separados.append(unidad)
                    
                    self.progress.emit(int((ll / Total) * 100)+1)  # Emitir progreso
                    
            # self.progress.emit(100)
            
            
            Data_Procesado.loc[len(Data_Procesado)] = [texto]+Separados[:len(columna)]
            
            if True:
                Data_Procesado = Data_Procesado.applymap(self.Numero)
                
            Agrupacion_Registros.append(Data_Procesado)

            ultima_agrup = texto[0]
        Tabla_Final = pd.concat(Agrupacion_Registros)
        Total_Registro.append(Tabla_Final)
        self.output_data = Total_Registro
        self.finished.emit()


class FileProcessorApp(QWidget):
    def __init__(self):
        super().__init__()
        self.archivo = "Formato_de_Registro_DJ_AT_2025.xlsx"
        workbook = load_workbook(filename= self.archivo )
        self.DJS = [x for x in workbook.sheetnames if x[0]=="F"]

        self.Total_Registro = []
        self.Data = []

        self.initUI()


    def initUI(self):
        self.setWindowTitle("DJ2TXT")
        self.setGeometry(100, 100, 400, 200)

        # Layout principal
        self.layout = QVBoxLayout()

        # Etiqueta para archivo
        
        self.barra_superior = QHBoxLayout()
        self.file_label = QLabel("Ningún archivo seleccionado")
        self.barra_superior.addWidget(self.file_label)
    
        self.combo_box = QComboBox()
        self.combo_box.addItems(self.DJS)
        self.barra_superior.addWidget(self.combo_box)
    
        self.layout.addLayout(self.barra_superior)

        # Botón para cargar archivo
        self.load_button = QPushButton("Cargar Archivo")
        self.load_button.clicked.connect(self.load_file)
        self.layout.addWidget(self.load_button)

        # Botón para procesar
        self.process_button = QPushButton("Procesar Archivo")
        self.process_button.setEnabled(False)
        self.process_button.clicked.connect(self.process_file)
        self.layout.addWidget(self.process_button)

        # Barra de progreso
        self.progress_bar = QProgressBar()
        self.progress_bar.setAlignment(Qt.AlignCenter)
        self.layout.addWidget(self.progress_bar)

        # Botón para exportar archivo
        self.export_button = QPushButton("Exportar Archivo")
        self.export_button.setEnabled(False)
        self.export_button.clicked.connect(self.export_file)
        self.layout.addWidget(self.export_button)

        # Mensaje de estado
        self.status_label = QLabel("Estado: Esperando acción")
        self.layout.addWidget(self.status_label)

        self.setLayout(self.layout)

    def load_file(self):
        options = QFileDialog.Options()
        file_name, _ = QFileDialog.getOpenFileName(self, "Seleccionar Archivo", "", "Todos los Archivos (*);;Archivos de Texto (*.txt)", options=options)
        if file_name:
            self.file_label.setText(f"Archivo: {file_name}")
            self.file_path = file_name
            self.process_button.setEnabled(True)

            with open(self.file_path, 'r') as archivo:
                
                for linea in archivo:
                    self.Data.append(linea.strip())
            self.status_label.setText("Estado: Archivo cargado")


    def numero(self,value):
        try:
            a = int(value)
            return True
        except:
            return False


    def Formatear_DJ(self):
        tabla = pd.read_excel(self.archivo,sheet_name=self.combo_box.currentText())
        tabla["Extensión"] = [
            int(re.findall(r'\d+', str(x))[0]) if "REGISTRO" in str(x) and re.findall(r'\d+', str(x)) else np.nan
            for x in tabla["ÍNDICE"]
        ]
        tabla["Extensión"] = tabla["Extensión"].fillna(method="ffill")
        tabla['Unnamed: 1'] = tabla['Unnamed: 1'].fillna(method='ffill')
        tabla = tabla.fillna("")
        tabla = tabla[tabla["Unnamed: 3"]!=""]
        tabla["Descripcion"]= [x1 + "_" + x2 if x2!="" else x1 for x1,x2 in zip(tabla["Unnamed: 1"],tabla["Unnamed: 2"])]
        tabla= tabla[[self.numero(x) for x in tabla["Unnamed: 3"].values]]

        tabla["Registro"]=tabla["Extensión"]
        tabla["Registro"] = tabla["Registro"].astype(int)
        tabla = tabla.drop(columns=["Extensión"])

        tabla = tabla.rename(columns={'ÍNDICE': 'Indice',
                                      'Unnamed: 3': 'Desde',
                                      'Unnamed: 4': 'Hasta',
                                      'Unnamed: 5': 'LARGO',
                                      })
        
        tabla_final = []
        for x in np.unique(tabla["Registro"].values):
            vol = tabla[tabla["Registro"] == x].copy()  # Se hace una copia para evitar advertencias
            
            if len(vol) > 1:
                comparacion = ["S"]
                for i in range(1, len(vol)):
                    if comparacion[-1] == "N":
                        comparacion.append("N")
                    else:
                        comparacion.append("S" if vol["Desde"].iloc[i] > vol["Desde"].iloc[i - 1] else "N")
                vol["Comparación"] = comparacion
            else:
                vol["Comparación"] = ["N"]
            
            vol = vol[vol["Comparación"]=="S"]
            tabla_final.append(vol.drop(columns=["Comparación"]))
        tabla = pd.concat(tabla_final)
        
        col = ["Indice","c1","C2","Desde","Hasta","LARGO","c3","c4","c5","Descripcion","Registro"]
        return tabla 


    def process_file(self):
        
        self.Formato = self.Formatear_DJ()
        self.progress_bar.setValue(0)
        self.status_label.setText("Estado: Procesando archivo...")
        self.worker = WorkerThread([self.Data, self.Formato])
        self.worker.progress.connect(self.progress_bar.setValue)
        self.worker.finished.connect(self.processing_complete)
        self.worker.start()


    @pyqtSlot()
    def processing_complete(self):
        self.status_label.setText("Estado: Procesamiento completado")
        self.Total_Registro = self.worker.output_data  # Recupera los datos del hilo
        self.export_button.setEnabled(True)
        

    def export_file(self):
        options = QFileDialog.Options()
        save_path, _ = QFileDialog.getSaveFileName(self, "Guardar Archivo", "", "Archivos Excel (*.xlsx)", options=options)
        if save_path:
            with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
                for i, x in enumerate(self.Total_Registro):
                    x.to_excel(writer, sheet_name=f'Registro {i}', index=False)
                    
            self.status_label.setText("Estado: Archivo exportado exitosamente")


if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = FileProcessorApp()
    ex.show()
    sys.exit(app.exec_())
